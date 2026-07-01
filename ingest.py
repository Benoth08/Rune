#!/usr/bin/env python3
"""Lythéa — ingestion de documents dans la mémoire long-terme.

Alimente la base ChromaDB de Lythéa avec le contenu de fichiers
(PDF, txt, md, docx). Une fois ingérés, ces documents deviennent
consultables par Lythéa via sa Phase B (RAG) — exactement comme les
souvenirs de conversations, mais persistants et thématiques.

Cas d'usage
-----------
- Spécialiser Lythéa sur le corpus documentaire d'une entreprise
  (procédures, rapports, specs techniques).
- Charger des articles scientifiques pour qu'elle aide à les
  analyser et croiser les informations.
- Constituer une base de connaissances de référence stable.

Usage
-----
    # Ingérer tous les documents d'un dossier
    python ingest.py /chemin/vers/mes_documents/

    # Ingérer un seul fichier
    python ingest.py rapport_2025.pdf

    # Purger les documents précédemment ingérés puis réingérer
    python ingest.py --reset /chemin/vers/mes_documents/

    # Tagger les documents avec une étiquette thématique
    python ingest.py --tag "physique_quantique" articles/

    # Lister ce qui est déjà ingéré
    python ingest.py --list

    # Supprimer tous les documents ingérés (sans réingérer)
    python ingest.py --purge

Notes
-----
- Le document est d'abord **assemblé en entier** (toutes les pages
  recollées), PUIS découpé en *chunks* (~800 caractères avec
  chevauchement). Cela garantit qu'une phrase coupée en bas de page
  n'est pas tronquée — elle est recollée avant le chunking.
- Les césures de ligne artificielles (fréquentes en extraction PDF)
  sont réparées : ``mot-\nsuite`` → ``motsuite``, et un ``\n`` en
  pleine phrase devient une espace. Les vraies ruptures (fin de
  phrase, paragraphes ``\n\n``) sont préservées.
- Chaque chunk garde tout de même son **numéro de page de début**
  via une table d'offsets — l'info de localisation n'est pas perdue.
- Chaque chunk porte une metadata ``type="knowledge"`` qui le
  distingue des souvenirs de conversation (``type="exchange"``).
- L'ingestion est idempotente par fichier : réingérer le même
  fichier remplace ses anciens chunks (détection par hash de
  contenu + nom).
- ChromaDB calcule lui-même les embeddings (all-MiniLM-L6-v2), donc
  aucun modèle Lythéa n'a besoin d'être chargé pour ingérer.
- Le BM25 sparse index de Lythéa se reconstruit automatiquement à
  la première requête après ingestion.
"""
from __future__ import annotations

import argparse
import hashlib
import sys
import time
from pathlib import Path

# ── Import ChromaDB + config Lythéa ────────────────────────────────────
try:
    import chromadb
except ImportError:
    print("❌ chromadb n'est pas installé.")
    print("   Installe-le avec : pip install chromadb --break-system-packages")
    sys.exit(1)

# Réutilise exactement le même chemin que l'app pour pointer sur la
# même base. Si l'import échoue (script lancé hors arborescence), on
# retombe sur un chemin par défaut compatible.
try:
    from rune.config import CHROMA_DIR
except Exception:
    from rune.env import CACHE_ROOT  # type: ignore
    CHROMA_DIR = CACHE_ROOT / "chroma"

COLLECTION_NAME = "lythea_memory"

# ── Paramètres de chunking ─────────────────────────────────────────────
CHUNK_SIZE = 800          # caractères par chunk (cible)
CHUNK_OVERLAP = 150       # chevauchement entre chunks consécutifs
MIN_CHUNK_SIZE = 80       # en-dessous, le chunk est ignoré (bruit)

SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".md", ".markdown", ".docx", ".rst"}


# ═══════════════════════════════════════════════════════════════════════
# Extraction de texte par type de fichier
# ═══════════════════════════════════════════════════════════════════════

def extract_pdf(path: Path) -> list[tuple[int, str]]:
    """Extrait le texte d'un PDF, page par page.

    Returns
    -------
    list[tuple[int, str]]
        Liste de ``(numéro_page, texte_page)``.
    """
    try:
        import pdfplumber
    except ImportError:
        print("❌ pdfplumber requis pour les PDF.")
        print("   pip install pdfplumber --break-system-packages")
        sys.exit(1)

    pages: list[tuple[int, str]] = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            text = page.extract_text() or ""
            if text.strip():
                pages.append((i, text))
    return pages


def extract_docx(path: Path) -> list[tuple[int, str]]:
    """Extrait le texte d'un .docx (paragraphes concaténés).

    Pas de notion de page dans docx → on renvoie tout en 'page 1'.
    """
    try:
        import docx  # python-docx
    except ImportError:
        print("❌ python-docx requis pour les .docx.")
        print("   pip install python-docx --break-system-packages")
        sys.exit(1)

    document = docx.Document(str(path))
    paragraphs = [p.text for p in document.paragraphs if p.text.strip()]
    full_text = "\n".join(paragraphs)
    return [(1, full_text)] if full_text.strip() else []


def extract_xlsx(path: Path) -> list[tuple[int, str]]:
    """V6.0.0-rc rev5 — Extrait le texte d'un classeur Excel (.xlsx).

    Stratégie :
    - Une "page" = une feuille. Permet à Lythéa de référencer
      « feuille 2 » naturellement, et facilite le chunking si une
      feuille déborde.
    - Chaque feuille est rendue en TSV (séparateur tabulation), plus
      lisible pour le LLM qu'un CSV avec virgules dans les cellules.
    - Les cellules vides sont représentées par "" (chaîne vide), pas
      par "None" — évite la pollution textuelle.
    - On ignore les feuilles cachées et les cellules formule (on
      prend la valeur calculée si disponible, sinon la formule).

    Limites :
    - Pas de mise en forme (couleur, gras) — texte brut uniquement.
    - Les graphiques et images sont ignorés.
    - Pour les très gros fichiers (>100k cellules), on truncate.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("❌ openpyxl requis pour les .xlsx.")
        print("   pip install openpyxl --break-system-packages")
        sys.exit(1)

    # data_only=True : prendre la VALEUR calculée des formules, pas la
    # formule elle-même. Plus utile pour le LLM (il veut le résultat).
    wb = load_workbook(str(path), data_only=True, read_only=True)

    pages: list[tuple[int, str]] = []
    page_idx = 0
    MAX_CELLS_PER_SHEET = 10_000  # garde-fou taille

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Skip hidden sheets (workflow d'admin / temp)
        if ws.sheet_state in ("hidden", "veryHidden"):
            continue

        page_idx += 1
        lines = [f"# Feuille : {sheet_name}"]
        n_cells = 0
        truncated = False

        for row in ws.iter_rows(values_only=True):
            # Skip lignes complètement vides
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            # Rendre chaque cellule comme str ; None → ""
            cells = [str(c) if c is not None else "" for c in row]
            lines.append("\t".join(cells))
            n_cells += len(row)
            if n_cells > MAX_CELLS_PER_SHEET:
                lines.append(
                    f"[… feuille tronquée à {MAX_CELLS_PER_SHEET} cellules]"
                )
                truncated = True
                break

        sheet_text = "\n".join(lines)
        if sheet_text.strip():
            pages.append((page_idx, sheet_text))

    wb.close()
    return pages


def extract_plaintext(path: Path) -> list[tuple[int, str]]:
    """Extrait le texte d'un fichier texte brut (txt, md, rst…)."""
    text = path.read_text(encoding="utf-8", errors="replace")
    return [(1, text)] if text.strip() else []


def extract_html(path: Path) -> list[tuple[int, str]]:
    """Extrait le texte d'un fichier HTML.

    Utilise BeautifulSoup si disponible (rendu propre, sans tags).
    Fallback minimal sans dépendance : regex pour stripper les tags.
    Ce fallback est moins propre (espaces, entités HTML brutes) mais
    permet de lire le contenu si bs4 n'est pas installé.
    """
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(raw, "html.parser")
        # Retirer les scripts/styles qui polluent la sortie texte.
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text(separator="\n", strip=True)
    except ImportError:
        import re
        # Fallback : strip tags, décode quelques entités courantes.
        text = re.sub(r"<script[^>]*>.*?</script>", "", raw,
                      flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text,
                      flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<[^>]+>", " ", text)
        text = (text.replace("&nbsp;", " ").replace("&amp;", "&")
                    .replace("&lt;", "<").replace("&gt;", ">")
                    .replace("&quot;", '"').replace("&#39;", "'"))
        # Compacter les whitespaces.
        text = re.sub(r"[ \t]+", " ", text)
        text = re.sub(r"\n\s*\n+", "\n\n", text)
        text = text.strip()
    return [(1, text)] if text.strip() else []


def extract_csv(path: Path) -> list[tuple[int, str]]:
    """Convertit un CSV en texte lisible (header : valeur1 | valeur2…)."""
    import csv
    lines: list[str] = []
    try:
        with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
            # Sniff le délimiteur (virgule, point-virgule, tab).
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel  # défaut : virgule
            reader = csv.reader(fh, dialect)
            header: list[str] | None = None
            for i, row in enumerate(reader):
                if i == 0 and any(c.strip() for c in row):
                    header = [c.strip() for c in row]
                    lines.append(" | ".join(header))
                    lines.append("-" * 40)
                else:
                    cells = [c.strip() for c in row]
                    if header and len(cells) == len(header):
                        lines.append("\n".join(
                            f"{h}: {v}" for h, v in zip(header, cells) if v
                        ))
                        lines.append("")
                    else:
                        lines.append(" | ".join(cells))
    except Exception as exc:
        return [(1, f"[Erreur lecture CSV: {exc}]")]
    text = "\n".join(lines).strip()
    return [(1, text)] if text else []


def extract_json(path: Path) -> list[tuple[int, str]]:
    """Lit un JSON et le re-formate proprement (indent=2)."""
    import json
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
        data = json.loads(raw)
        text = json.dumps(data, indent=2, ensure_ascii=False)
    except Exception as exc:
        # Si invalide, on renvoie le texte brut.
        text = path.read_text(encoding="utf-8", errors="replace")
        text = f"[Note : JSON invalide ({exc}), texte brut ci-dessous]\n\n{text}"
    return [(1, text)] if text.strip() else []


def extract_xml(path: Path) -> list[tuple[int, str]]:
    """Extrait le texte d'un XML en aplatissant la structure.

    On retire les balises mais on garde le texte de chaque nœud,
    séparé par des sauts de ligne. Pour la cohérence avec HTML on
    réutilise BeautifulSoup si dispo (parseur XML).
    """
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(raw, "xml")
        text = soup.get_text(separator="\n", strip=True)
    except ImportError:
        import re
        text = re.sub(r"<[^>]+>", "\n", raw)
        text = re.sub(r"\n\s*\n+", "\n\n", text).strip()
    return [(1, text)] if text.strip() else []


def extract_text(path: Path) -> list[tuple[int, str]]:
    """Dispatch d'extraction selon l'extension du fichier.

    Formats supportés :
    - PDF (pdfplumber)
    - DOCX (python-docx)
    - XLSX (openpyxl)  [V6.0.0-rc rev5]
    - Texte brut : .txt, .md, .markdown, .rst
    - HTML : .html, .htm (BeautifulSoup ou fallback regex)
    - Données : .csv (csv stdlib), .json (json stdlib), .xml (BeautifulSoup ou regex)
    """
    ext = path.suffix.lower()
    if ext == ".pdf":
        return extract_pdf(path)
    if ext == ".docx":
        return extract_docx(path)
    if ext == ".xlsx":
        return extract_xlsx(path)
    if ext in {".txt", ".md", ".markdown", ".rst"}:
        return extract_plaintext(path)
    if ext in {".html", ".htm"}:
        return extract_html(path)
    if ext == ".csv":
        return extract_csv(path)
    if ext == ".json":
        return extract_json(path)
    if ext == ".xml":
        return extract_xml(path)
    raise ValueError(f"Extension non supportée : {ext}")


# ═══════════════════════════════════════════════════════════════════════
# Chunking
# ═══════════════════════════════════════════════════════════════════════

def chunk_text(text: str, size: int = CHUNK_SIZE,
               overlap: int = CHUNK_OVERLAP) -> list[tuple[str, int]]:
    """Découpe un texte en chunks avec chevauchement.

    La stratégie privilégie les coupures sur les frontières naturelles
    (paragraphe, puis phrase, puis espace) pour ne pas casser le sens
    en plein milieu d'un mot ou d'une idée.

    Returns
    -------
    list[tuple[str, int]]
        Liste de ``(texte_du_chunk, position_de_départ)``. La position
        est l'offset (en caractères) du chunk dans le ``text`` d'entrée
        — nécessaire pour retrouver à quelle page il appartient via
        ``page_for_offset``.
    """
    text = text.strip()
    if len(text) <= size:
        return [(text, 0)] if len(text) >= MIN_CHUNK_SIZE else []

    chunks: list[tuple[str, int]] = []
    start = 0
    n = len(text)

    while start < n:
        end = min(start + size, n)

        # Si on n'est pas à la fin, essayer de couper proprement.
        if end < n:
            # Chercher une frontière de paragraphe dans la dernière
            # moitié du chunk.
            window_start = start + size // 2
            para_break = text.rfind("\n\n", window_start, end)
            if para_break > window_start:
                end = para_break
            else:
                # Sinon une fin de phrase.
                sentence_break = max(
                    text.rfind(". ", window_start, end),
                    text.rfind("! ", window_start, end),
                    text.rfind("? ", window_start, end),
                    text.rfind(".\n", window_start, end),
                )
                if sentence_break > window_start:
                    end = sentence_break + 1
                else:
                    # En dernier recours, couper sur un espace.
                    space_break = text.rfind(" ", window_start, end)
                    if space_break > window_start:
                        end = space_break

        # Capturer l'offset AVANT le strip — le strip ne décale le
        # début que de quelques espaces, négligeable pour le mapping
        # de page, et on garde ainsi une position cohérente avec le
        # texte source.
        chunk_start = start
        chunk = text[start:end].strip()
        if len(chunk) >= MIN_CHUNK_SIZE:
            chunks.append((chunk, chunk_start))

        # Avancer en gardant le chevauchement.
        next_start = end - overlap
        # Garde-fou : toujours progresser d'au moins 1 caractère.
        start = next_start if next_start > start else end

    return chunks


# ═══════════════════════════════════════════════════════════════════════
# Identifiants stables
# ═══════════════════════════════════════════════════════════════════════

def file_fingerprint(path: Path) -> str:
    """Hash court et stable d'un fichier (nom + contenu).

    Sert à générer des ids de chunks déterministes : réingérer le
    même fichier produit les mêmes ids → remplacement propre via
    ``upsert`` plutôt que doublons.
    """
    h = hashlib.sha256()
    h.update(path.name.encode("utf-8"))
    h.update(path.read_bytes())
    return h.hexdigest()[:16]


# ═══════════════════════════════════════════════════════════════════════
# Assemblage document : recoller les pages avant de chunker
# ═══════════════════════════════════════════════════════════════════════

def _dehyphenate_and_join(text: str) -> str:
    """Recolle les césures artificielles à l'intérieur d'un bloc de texte.

    Les extracteurs PDF insèrent un ``\\n`` à chaque fin de ligne
    visuelle. Quand une phrase continue sur la ligne suivante, ce
    ``\\n`` est artificiel et casse la phrase. Heuristique :

    - ``mot-\\nsuite``  → ``motsuite``   (césure typographique : trait
      d'union en fin de ligne, on recolle sans espace)
    - ``mot\\nsuite``   → ``mot suite``  (la ligne ne finit pas par une
      ponctuation de fin de phrase → continuation, on met un espace)
    - ``phrase.\\nAutre`` → inchangé    (ponctuation de fin → vraie
      rupture, le ``\\n`` est conservé)
    - ``\\n\\n``        → inchangé       (paragraphe : structure réelle)

    Ce nettoyage est volontairement conservateur : en cas de doute on
    préfère garder le ``\\n``. L'objectif est de réparer les césures
    évidentes, pas de reformater le document.
    """
    import re

    # 1. Césure typographique : "mot-\n" → "mot" (recoller le mot).
    text = re.sub(r"-\n(?=\w)", "", text)

    # 2. Continuation de ligne : un \n simple précédé d'un caractère
    #    qui n'est PAS une fin de phrase (. ! ? : ; » ) ]) et suivi
    #    d'une minuscule ou d'un chiffre → c'est une continuation,
    #    on remplace par une espace. Le (?<!\n) et (?!\n) garantissent
    #    qu'on ne touche pas aux doubles newlines (paragraphes).
    text = re.sub(
        r"(?<![\.\!\?:;»\)\]\n])\n(?!\n)(?=[a-zà-öø-ÿ0-9])",
        " ",
        text,
    )

    return text


def assemble_document(
    pages: list[tuple[int, str]],
) -> tuple[str, list[tuple[int, int]]]:
    """Concatène les pages en un seul texte continu.

    C'est l'étape clé : chunker page par page coupe les phrases qui
    chevauchent une frontière de page. En recollant tout le document
    AVANT de chunker, ces phrases restent entières.

    Pour ne pas perdre l'information de page, on construit en parallèle
    une table d'offsets ``[(char_start, page_num), ...]`` : à quelle
    position dans le texte assemblé commence chaque page. ``page_for_offset``
    s'en sert ensuite pour retrouver la page d'un chunk donné.

    Chaque page passe d'abord par ``_dehyphenate_and_join`` qui répare
    les césures de ligne artificielles (fréquentes en extraction PDF).
    Le séparateur entre pages est ensuite choisi dynamiquement :

    - si la page précédente se termine par une ponctuation de fin de
      phrase → ``\\n`` (rupture nette, page complète)
    - sinon → une espace : la phrase courait d'une page à l'autre, on
      la recolle proprement.

    Returns
    -------
    tuple[str, list[tuple[int, int]]]
        ``(texte_complet, table_offsets)``.
    """
    # Ponctuations qui signalent une vraie fin de phrase / bloc.
    _SENTENCE_END = (".", "!", "?", ":", "»", ")", "]", '"')

    parts: list[str] = []
    offsets: list[tuple[int, int]] = []
    cursor = 0

    for page_num, page_text in pages:
        page_text = _dehyphenate_and_join(page_text.strip())
        if not page_text:
            continue

        # Choisir le séparateur AVEC la page précédente.
        if parts:
            prev_tail = parts[-1].rstrip()
            sep = "\n" if prev_tail.endswith(_SENTENCE_END) else " "
        else:
            sep = ""

        if sep:
            # Le séparateur compte dans les offsets.
            cursor += len(sep)
        offsets.append((cursor, page_num))
        parts.append(sep + page_text if sep else page_text)
        cursor += len(page_text)

    full_text = "".join(parts)
    return full_text, offsets


def page_for_offset(offset: int, offsets: list[tuple[int, int]]) -> int:
    """Retrouve le numéro de page contenant une position de caractère.

    ``offsets`` est la table ``[(char_start, page_num), ...]`` produite
    par ``assemble_document``, triée par ``char_start`` croissant. On
    cherche la dernière page dont le ``char_start`` est <= offset.

    Recherche linéaire : les documents font rarement plus de quelques
    centaines de pages, inutile de sortir la dichotomie.
    """
    page = offsets[0][1] if offsets else 1
    for char_start, page_num in offsets:
        if char_start <= offset:
            page = page_num
        else:
            break
    return page


# ═══════════════════════════════════════════════════════════════════════
# Opérations ChromaDB
# ═══════════════════════════════════════════════════════════════════════

def get_collection():
    """Ouvre (ou crée) la collection ChromaDB de Lythéa."""
    CHROMA_DIR.mkdir(parents=True, exist_ok=True)
    client = chromadb.PersistentClient(
        path=str(CHROMA_DIR),
        settings=chromadb.Settings(anonymized_telemetry=False),
    )
    return client.get_or_create_collection(COLLECTION_NAME)


def remove_file_chunks(collection, source_name: str) -> int:
    """Supprime tous les chunks d'un fichier source donné.

    Returns
    -------
    int
        Nombre de chunks supprimés.
    """
    try:
        existing = collection.get(
            where={"source": source_name},
            include=[],
        )
        ids = existing.get("ids", [])
        if ids:
            collection.delete(ids=ids)
        return len(ids)
    except Exception as exc:
        print(f"   ⚠️  Nettoyage de '{source_name}' échoué : {exc}")
        return 0


def ingest_file(collection, path: Path, tag: str | None) -> tuple[int, int]:
    """Ingère un fichier dans ChromaDB.

    Returns
    -------
    tuple[int, int]
        ``(chunks_ajoutés, chunks_remplacés)``.
    """
    source_name = path.name

    # Remplacement propre : on purge les anciens chunks de ce fichier.
    replaced = remove_file_chunks(collection, source_name)

    try:
        pages = extract_text(path)
    except Exception as exc:
        print(f"   ❌ Extraction échouée : {exc}")
        return (0, replaced)

    if not pages:
        print(f"   ⚠️  Aucun texte extrait (fichier vide ou scanné ?)")
        return (0, replaced)

    fingerprint = file_fingerprint(path)
    now = time.time()

    # ── Assembler tout le document AVANT de chunker ──────────────────
    # Chunker page par page couperait les phrases qui chevauchent une
    # frontière de page. On recolle d'abord, puis on chunke le texte
    # continu — les phrases coupées se reconstituent. La table
    # d'offsets permet quand même de retrouver la page de chaque chunk.
    full_text, page_offsets = assemble_document(pages)

    if not full_text.strip():
        print(f"   ⚠️  Aucun texte exploitable après assemblage")
        return (0, replaced)

    documents: list[str] = []
    ids: list[str] = []
    metadatas: list[dict] = []

    chunk_counter = 0
    for chunk, char_start in chunk_text(full_text):
        chunk_counter += 1
        chunk_id = f"kb_{fingerprint}_{chunk_counter:04d}"
        # Retrouver la page où commence ce chunk. Un chunk peut
        # chevaucher deux pages — on retient la page de DÉBUT, c'est
        # la convention la plus intuitive ("ce passage commence p.12").
        page_num = page_for_offset(char_start, page_offsets)
        meta = {
            "type": "knowledge",       # distinct des "exchange"
            "source": source_name,
            "page": page_num,
            "chunk": chunk_counter,
            "ingested_at": now,
            "fingerprint": fingerprint,
        }
        if tag:
            meta["tag"] = tag
        documents.append(chunk)
        ids.append(chunk_id)
        metadatas.append(meta)

    if not documents:
        print(f"   ⚠️  Aucun chunk valide produit")
        return (0, replaced)

    # Ajout par batch (ChromaDB gère bien jusqu'à ~5000 docs/batch ;
    # on découpe par sécurité sur les très gros fichiers).
    BATCH = 500
    for i in range(0, len(documents), BATCH):
        collection.add(
            documents=documents[i:i + BATCH],
            ids=ids[i:i + BATCH],
            metadatas=metadatas[i:i + BATCH],
        )

    return (len(documents), replaced)


def list_knowledge(collection) -> None:
    """Affiche un récapitulatif des documents de connaissance ingérés."""
    try:
        data = collection.get(
            where={"type": "knowledge"},
            include=["metadatas"],
        )
    except Exception as exc:
        print(f"❌ Lecture de la collection échouée : {exc}")
        return

    metas = data.get("metadatas", [])
    if not metas:
        print("📭 Aucun document de connaissance ingéré pour le moment.")
        return

    # Agréger par fichier source.
    by_source: dict[str, dict] = {}
    for m in metas:
        src = m.get("source", "?")
        if src not in by_source:
            by_source[src] = {
                "chunks": 0,
                "tag": m.get("tag", ""),
                "ingested_at": m.get("ingested_at", 0),
            }
        by_source[src]["chunks"] += 1

    print(f"\n📚 {len(by_source)} document(s) ingéré(s), "
          f"{len(metas)} chunk(s) au total :\n")
    for src, info in sorted(by_source.items()):
        tag_str = f"  [tag: {info['tag']}]" if info["tag"] else ""
        when = ""
        if info["ingested_at"]:
            when = time.strftime(
                "%Y-%m-%d %H:%M", time.localtime(info["ingested_at"])
            )
        print(f"  • {src}")
        print(f"      {info['chunks']} chunks{tag_str}  —  ingéré le {when}")
    print()


def purge_knowledge(collection) -> int:
    """Supprime TOUS les documents de connaissance (type=knowledge).

    Ne touche pas aux souvenirs de conversation (type=exchange).

    Returns
    -------
    int
        Nombre de chunks supprimés.
    """
    try:
        data = collection.get(where={"type": "knowledge"}, include=[])
        ids = data.get("ids", [])
        if ids:
            collection.delete(ids=ids)
        return len(ids)
    except Exception as exc:
        print(f"❌ Purge échouée : {exc}")
        return 0


# ═══════════════════════════════════════════════════════════════════════
# Collecte des fichiers
# ═══════════════════════════════════════════════════════════════════════

def collect_files(target: Path) -> list[Path]:
    """Liste les fichiers ingérables à partir d'un chemin (fichier ou dossier)."""
    if target.is_file():
        if target.suffix.lower() in SUPPORTED_EXTENSIONS:
            return [target]
        print(f"⚠️  Extension non supportée : {target.suffix}")
        return []

    if target.is_dir():
        files = sorted(
            p for p in target.rglob("*")
            if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
        )
        return files

    print(f"❌ Chemin introuvable : {target}")
    return []


# ═══════════════════════════════════════════════════════════════════════
# Entrée principale
# ═══════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Ingestion de documents dans la mémoire long-terme de Lythéa.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Exemples :\n"
            "  python ingest.py documents/\n"
            "  python ingest.py --tag physique_quantique articles/\n"
            "  python ingest.py --reset corpus_entreprise/\n"
            "  python ingest.py --list\n"
            "  python ingest.py --purge\n"
        ),
    )
    parser.add_argument(
        "path", nargs="?", default=None,
        help="Fichier ou dossier à ingérer (PDF, txt, md, docx, rst).",
    )
    parser.add_argument(
        "--tag", default=None,
        help="Étiquette thématique attachée aux documents (ex: 'rh', 'r_and_d').",
    )
    parser.add_argument(
        "--reset", action="store_true",
        help="Purge tous les documents de connaissance avant d'ingérer.",
    )
    parser.add_argument(
        "--list", action="store_true",
        help="Liste les documents déjà ingérés et quitte.",
    )
    parser.add_argument(
        "--purge", action="store_true",
        help="Supprime tous les documents de connaissance et quitte.",
    )
    args = parser.parse_args()

    print("═" * 60)
    print("  Lythéa — Ingestion de documents")
    print(f"  Base ChromaDB : {CHROMA_DIR}")
    print("═" * 60)

    collection = get_collection()

    # ── Modes informatifs / destructifs sans ingestion ────────────────
    if args.list:
        list_knowledge(collection)
        return

    if args.purge:
        n = purge_knowledge(collection)
        print(f"\n🗑️  {n} chunk(s) de connaissance supprimé(s).")
        print("    (Les souvenirs de conversation sont intacts.)\n")
        return

    # ── Mode ingestion ────────────────────────────────────────────────
    if not args.path:
        parser.print_help()
        print("\n❌ Indique un fichier ou un dossier à ingérer.")
        sys.exit(1)

    if args.reset:
        n = purge_knowledge(collection)
        print(f"\n🗑️  --reset : {n} ancien(s) chunk(s) supprimé(s).\n")

    target = Path(args.path).expanduser().resolve()
    files = collect_files(target)

    if not files:
        print("\n❌ Aucun fichier ingérable trouvé.")
        print(f"   Extensions supportées : {', '.join(sorted(SUPPORTED_EXTENSIONS))}")
        sys.exit(1)

    print(f"\n📂 {len(files)} fichier(s) à ingérer")
    if args.tag:
        print(f"🏷️  Tag : {args.tag}")
    print()

    total_added = 0
    total_replaced = 0
    failed = 0
    t0 = time.time()

    for idx, path in enumerate(files, start=1):
        # Chemin relatif lisible si possible.
        try:
            display = path.relative_to(target if target.is_dir() else target.parent)
        except ValueError:
            display = path.name
        print(f"[{idx}/{len(files)}] {display}")

        try:
            added, replaced = ingest_file(collection, path, args.tag)
            total_added += added
            total_replaced += replaced
            if added:
                msg = f"   ✅ {added} chunks ajoutés"
                if replaced:
                    msg += f" ({replaced} remplacés)"
                print(msg)
            else:
                failed += 1
        except KeyboardInterrupt:
            print("\n\n⚠️  Interrompu par l'utilisateur.")
            break
        except Exception as exc:
            print(f"   ❌ Erreur : {exc}")
            failed += 1

    elapsed = time.time() - t0
    print()
    print("═" * 60)
    print(f"  Terminé en {elapsed:.1f}s")
    print(f"  {total_added} chunks ajoutés"
          + (f", {total_replaced} remplacés" if total_replaced else "")
          + (f", {failed} fichier(s) en échec" if failed else ""))
    final_count = collection.count()
    print(f"  Total dans la base : {final_count} documents "
          f"(connaissance + conversations)")
    print("═" * 60)
    print()
    print("💡 Lythéa exploitera ces documents dès sa prochaine requête.")
    print("   L'index BM25 se reconstruit automatiquement — rien à faire.")
    print()


if __name__ == "__main__":
    main()
